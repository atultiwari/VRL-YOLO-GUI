"""Task-aware report generators for batch inference output.

Three output formats, one task-shape input. Each generator receives a
list of `ReportItem` records that the routing layer already validated;
this module just turns them into bytes for download.

Why ReportLab + OpenPyXL here instead of, say, a Jinja template + a
headless browser-to-PDF? Three reasons:

- Both libraries are already pinned in `pyproject.toml`'s base deps
  (cheaper bundle than carrying Chromium).
- They generate fully synchronous output, which lets the FastAPI route
  return a `Response(content=bytes, ...)` in one go instead of streaming
  a chunked response (simpler to wire, simpler to download from JS).
- The clinical PDF template is intentionally low-fidelity (cover +
  summary + table + thumbnail grid) — Platypus' built-in Flowables are
  plenty for that, and adding Jinja would just hide what's getting
  rendered.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    Image as RLImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

Task = Literal["detect", "classify"]


# --- Input dataclasses ------------------------------------------------


@dataclass(frozen=True)
class DetectBoxRecord:
    class_name: str
    conf: float


@dataclass(frozen=True)
class ClassifyPredictionRecord:
    class_name: str
    conf: float


@dataclass(frozen=True)
class ReportItem:
    """One row in the per-image table.

    For detect items, `boxes` and `counts_per_class` are populated and
    the classify-only fields stay None. The router enforces this on the
    way in so the generators don't have to second-guess.
    """

    filename: str
    inference_ms: float
    # Detect
    boxes: tuple[DetectBoxRecord, ...] | None = None
    counts_per_class: dict[str, int] | None = None
    # Classify
    top1: ClassifyPredictionRecord | None = None
    top5: tuple[ClassifyPredictionRecord, ...] | None = None
    # PDF only: full image bytes for the embed grid (limited to a few
    # representative samples by the caller — we don't dictate which).
    image_bytes: bytes | None = None


@dataclass(frozen=True)
class ReportPayload:
    task: Task
    model: str
    items: tuple[ReportItem, ...]
    review_threshold: float
    # Optional aggregates passed from the frontend (saves the report
    # generator from re-rolling them). None means "compute it here".
    detect_per_class: dict[str, int] | None = None
    classify_per_class: dict[str, int] | None = None
    classify_flagged_count: int | None = None


# --- CSV --------------------------------------------------------------


def generate_csv(payload: ReportPayload) -> bytes:
    """One row per image. Columns differ by task."""
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)

    if payload.task == "detect":
        # Stable column order: filename, total boxes, top class, max conf,
        # ms, then one column per class (sorted by total descending).
        class_totals = _detect_class_totals(payload)
        classes = [c for c, _ in class_totals]
        writer.writerow(
            ["filename", "total_boxes", "top_class", "max_conf", "inference_ms", *classes]
        )
        for item in payload.items:
            boxes = item.boxes or ()
            counts = item.counts_per_class or {}
            total = len(boxes)
            max_conf = max((b.conf for b in boxes), default=0.0)
            top_class = (
                max(counts.items(), key=lambda kv: kv[1])[0] if counts else ""
            )
            per_class_counts = [counts.get(c, 0) for c in classes]
            writer.writerow(
                [
                    item.filename,
                    total,
                    top_class,
                    f"{max_conf:.4f}",
                    f"{item.inference_ms:.1f}",
                    *per_class_counts,
                ]
            )
    else:
        # Classify columns: filename, top1, top1_conf, then top2..top5
        # name + conf, "needs_review" flag (vs payload.review_threshold).
        writer.writerow(
            [
                "filename",
                "top1",
                "top1_conf",
                "top2",
                "top2_conf",
                "top3",
                "top3_conf",
                "top4",
                "top4_conf",
                "top5",
                "top5_conf",
                "needs_review",
                "inference_ms",
            ]
        )
        for item in payload.items:
            top5 = list(item.top5 or ())
            while len(top5) < 5:
                top5.append(ClassifyPredictionRecord(class_name="", conf=0.0))
            row: list[str | int | float] = [item.filename]
            for p in top5[:5]:
                row.append(p.class_name)
                row.append(f"{p.conf:.4f}")
            top1_conf = (item.top1.conf if item.top1 else 0.0)
            row.append(int(top1_conf < payload.review_threshold))
            row.append(f"{item.inference_ms:.1f}")
            writer.writerow(row)

    return buf.getvalue().encode("utf-8")


# --- XLSX -------------------------------------------------------------


_HEADER_FILL = PatternFill("solid", fgColor="FFE0E7FF")
_HEADER_FONT = Font(bold=True, color="FF1E293B")
_REVIEW_FILL = PatternFill("solid", fgColor="FFFFF4D6")


def generate_xlsx(payload: ReportPayload) -> bytes:
    """Two-sheet workbook: per-image table + aggregate summary."""
    wb = Workbook()
    per_image = wb.active
    per_image.title = "Per image"
    _write_xlsx_per_image(per_image, payload)

    aggregate_sheet = wb.create_sheet("Aggregate")
    _write_xlsx_aggregate(aggregate_sheet, payload)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _write_xlsx_per_image(ws, payload: ReportPayload) -> None:  # type: ignore[no-untyped-def]
    if payload.task == "detect":
        class_totals = _detect_class_totals(payload)
        classes = [c for c, _ in class_totals]
        headers = [
            "Filename",
            "Total boxes",
            "Top class",
            "Max conf",
            "Inference ms",
            *classes,
        ]
        _write_xlsx_header(ws, headers)
        for row_i, item in enumerate(payload.items, start=2):
            boxes = item.boxes or ()
            counts = item.counts_per_class or {}
            total = len(boxes)
            max_conf = max((b.conf for b in boxes), default=0.0)
            top_class = (
                max(counts.items(), key=lambda kv: kv[1])[0] if counts else ""
            )
            ws.cell(row=row_i, column=1, value=item.filename)
            ws.cell(row=row_i, column=2, value=total)
            ws.cell(row=row_i, column=3, value=top_class)
            ws.cell(row=row_i, column=4, value=round(max_conf, 4))
            ws.cell(row=row_i, column=5, value=round(item.inference_ms, 1))
            for col_i, cls in enumerate(classes, start=6):
                ws.cell(row=row_i, column=col_i, value=counts.get(cls, 0))
    else:
        headers = [
            "Filename",
            "Top-1 class",
            "Top-1 conf",
            "Top-2 class",
            "Top-2 conf",
            "Top-3 class",
            "Top-3 conf",
            "Top-4 class",
            "Top-4 conf",
            "Top-5 class",
            "Top-5 conf",
            "Needs review",
            "Inference ms",
        ]
        _write_xlsx_header(ws, headers)
        for row_i, item in enumerate(payload.items, start=2):
            top5 = list(item.top5 or ())
            while len(top5) < 5:
                top5.append(ClassifyPredictionRecord(class_name="", conf=0.0))
            top1_conf = item.top1.conf if item.top1 else 0.0
            needs_review = top1_conf < payload.review_threshold

            ws.cell(row=row_i, column=1, value=item.filename)
            for k in range(5):
                ws.cell(row=row_i, column=2 + k * 2, value=top5[k].class_name)
                ws.cell(row=row_i, column=3 + k * 2, value=round(top5[k].conf, 4))
            ws.cell(row=row_i, column=12, value="yes" if needs_review else "no")
            ws.cell(row=row_i, column=13, value=round(item.inference_ms, 1))

            if needs_review:
                for col in range(1, 14):
                    ws.cell(row=row_i, column=col).fill = _REVIEW_FILL

    _autosize_columns(ws)


def _write_xlsx_aggregate(ws, payload: ReportPayload) -> None:  # type: ignore[no-untyped-def]
    ws.cell(row=1, column=1, value="VRL YOLO GUI — Aggregate").font = Font(
        bold=True, size=14
    )
    ws.cell(
        row=2,
        column=1,
        value=f"Model: {payload.model}  ·  Task: {payload.task}  ·  Items: {len(payload.items)}",
    )
    ws.cell(
        row=3,
        column=1,
        value=f"Generated: {_now_iso()}",
    )
    if payload.task == "detect":
        ws.cell(row=5, column=1, value="Class").font = _HEADER_FONT
        ws.cell(row=5, column=2, value="Total boxes").font = _HEADER_FONT
        ws.cell(row=5, column=3, value="Max conf").font = _HEADER_FONT
        for cell in ws[5][:3]:
            cell.fill = _HEADER_FILL
        class_totals = _detect_class_totals(payload)
        for i, (cls, total) in enumerate(class_totals, start=6):
            ws.cell(row=i, column=1, value=cls)
            ws.cell(row=i, column=2, value=total)
            max_conf = max(
                (
                    b.conf
                    for item in payload.items
                    for b in (item.boxes or ())
                    if b.class_name == cls
                ),
                default=0.0,
            )
            ws.cell(row=i, column=3, value=round(max_conf, 4))
    else:
        ws.cell(row=5, column=1, value="Top-1 class").font = _HEADER_FONT
        ws.cell(row=5, column=2, value="Images").font = _HEADER_FONT
        ws.cell(row=5, column=3, value="Mean conf").font = _HEADER_FONT
        for cell in ws[5][:3]:
            cell.fill = _HEADER_FILL
        class_totals = _classify_top1_distribution(payload)
        for i, (cls, count, mean_conf) in enumerate(class_totals, start=6):
            ws.cell(row=i, column=1, value=cls)
            ws.cell(row=i, column=2, value=count)
            ws.cell(row=i, column=3, value=round(mean_conf, 4))

    _autosize_columns(ws)


def _write_xlsx_header(ws, headers: list[str]) -> None:  # type: ignore[no-untyped-def]
    for col_i, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")


def _autosize_columns(ws) -> None:  # type: ignore[no-untyped-def]
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


# --- PDF --------------------------------------------------------------


def generate_pdf(payload: ReportPayload) -> bytes:
    """Cover + summary + per-class breakdown + sample-image grid + table.

    Designed for screen and 1-page-per-section A4 printing. Images
    embedded as Flowables in their own section; PDFs are kept under a
    few MB even with ~20 sample patches.
    """
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=20 * mm,
        title=f"VRL YOLO GUI report — {payload.model}",
        author="VRL YOLO GUI",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle(
        "h1",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0a2540"),
        spaceAfter=4,
    )
    sub = ParagraphStyle(
        "sub",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#475569"),
    )
    h2 = ParagraphStyle(
        "h2",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#0a2540"),
        spaceBefore=10,
        spaceAfter=6,
    )
    footer = ParagraphStyle(
        "footer",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#94a3b8"),
        alignment=1,
    )

    story: list = []  # type: ignore[type-arg]
    title_text = (
        "Detection report"
        if payload.task == "detect"
        else "Classification report"
    )
    story.append(Paragraph(f"<b>{title_text}</b>", h1))
    story.append(
        Paragraph(
            f"Model: <b>{payload.model}</b> &nbsp;·&nbsp; "
            f"{len(payload.items)} image{'s' if len(payload.items) != 1 else ''} processed "
            f"&nbsp;·&nbsp; Review threshold {payload.review_threshold:.2f}",
            sub,
        )
    )
    story.append(Paragraph(f"Generated {_now_iso()}", sub))
    story.append(Spacer(1, 6 * mm))

    # --- Aggregate summary table ---
    story.append(Paragraph("Aggregate", h2))
    if payload.task == "detect":
        class_totals = _detect_class_totals(payload)
        data = [["Class", "Total boxes", "Max conf"]]
        for cls, total in class_totals:
            max_conf = max(
                (
                    b.conf
                    for item in payload.items
                    for b in (item.boxes or ())
                    if b.class_name == cls
                ),
                default=0.0,
            )
            data.append([cls, str(total), f"{max_conf * 100:.1f}%"])
    else:
        data = [["Top-1 class", "Images", "Mean conf"]]
        for cls, count, mean_conf in _classify_top1_distribution(payload):
            data.append([cls, str(count), f"{mean_conf * 100:.1f}%"])
    story.append(_build_table(data, header=True))

    # --- Sample image grid (if any image_bytes provided) ---
    samples_with_images = [it for it in payload.items if it.image_bytes]
    if samples_with_images:
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph("Sample images", h2))
        story.append(_image_grid(samples_with_images, sub))

    # --- Per-image table on its own page ---
    story.append(PageBreak())
    story.append(Paragraph("Per-image results", h2))
    story.append(_per_image_table(payload))

    # --- Footer note ---
    story.append(Spacer(1, 8 * mm))
    story.append(
        Paragraph(
            "Generated by VRL YOLO GUI · For research and demonstration use · "
            "Filenames may contain patient identifiers — anonymise before sharing.",
            footer,
        )
    )

    doc.build(story)
    return out.getvalue()


def _build_table(data: list[list[str]], *, header: bool) -> Table:
    table = Table(data, hAlign="LEFT", colWidths=None)
    style = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#cbd5e1")),
        ("LINEBELOW", (0, -1), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        style += [
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0a2540")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e7ff")),
        ]
    table.setStyle(TableStyle(style))
    return table


def _image_grid(items: list[ReportItem], caption_style: ParagraphStyle) -> Table:
    """3-column grid of thumbnails with filename captions."""
    cells: list[list[object]] = []
    row: list[object] = []
    cell_width = (170 / 3 - 4) * mm
    for idx, item in enumerate(items):
        try:
            img = RLImage(io.BytesIO(item.image_bytes or b""), width=cell_width, height=cell_width)
            img.hAlign = "CENTER"
        except Exception:  # noqa: BLE001 — corrupt sample shouldn't kill the PDF
            continue
        caption = Paragraph(item.filename, caption_style)
        from reportlab.platypus import KeepInFrame  # local import — only needed here
        block = KeepInFrame(cell_width, cell_width + 16, [img, caption])
        row.append(block)
        if len(row) == 3:
            cells.append(row)
            row = []
    if row:
        while len(row) < 3:
            row.append(Paragraph("", caption_style))
        cells.append(row)

    table = Table(cells, hAlign="LEFT", colWidths=[cell_width] * 3, rowHeights=None)
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6 * mm),
            ]
        )
    )
    return table


def _per_image_table(payload: ReportPayload) -> Table:
    if payload.task == "detect":
        rows: list[list[str]] = [["Filename", "Total", "Top class", "Max conf", "ms"]]
        for item in payload.items:
            boxes = item.boxes or ()
            counts = item.counts_per_class or {}
            top_class = (
                max(counts.items(), key=lambda kv: kv[1])[0] if counts else "—"
            )
            max_conf = max((b.conf for b in boxes), default=0.0)
            rows.append(
                [
                    item.filename,
                    str(len(boxes)),
                    top_class,
                    f"{max_conf * 100:.1f}%",
                    f"{item.inference_ms:.0f}",
                ]
            )
        return _build_table(rows, header=True)

    rows = [["Filename", "Top-1", "Conf", "Review", "ms"]]
    for item in payload.items:
        top1 = item.top1
        rows.append(
            [
                item.filename,
                top1.class_name if top1 else "—",
                f"{(top1.conf if top1 else 0.0) * 100:.1f}%",
                "yes" if (top1 and top1.conf < payload.review_threshold) else "",
                f"{item.inference_ms:.0f}",
            ]
        )
    return _build_table(rows, header=True)


# --- Aggregate helpers (shared by all formats) ------------------------


def _detect_class_totals(payload: ReportPayload) -> list[tuple[str, int]]:
    """Pre-computed totals from the frontend if available; else recount."""
    if payload.detect_per_class is not None:
        return sorted(payload.detect_per_class.items(), key=lambda kv: kv[1], reverse=True)
    counts: dict[str, int] = {}
    for item in payload.items:
        for cls, n in (item.counts_per_class or {}).items():
            counts[cls] = counts.get(cls, 0) + n
    return sorted(counts.items(), key=lambda kv: kv[1], reverse=True)


def _classify_top1_distribution(
    payload: ReportPayload,
) -> list[tuple[str, int, float]]:
    """List of (class_name, count, mean_conf) sorted by count descending."""
    acc: dict[str, tuple[int, float]] = {}
    for item in payload.items:
        if not item.top1:
            continue
        cur = acc.get(item.top1.class_name, (0, 0.0))
        acc[item.top1.class_name] = (cur[0] + 1, cur[1] + item.top1.conf)
    out = [
        (name, count, sum_conf / count) for name, (count, sum_conf) in acc.items()
    ]
    out.sort(key=lambda x: x[1], reverse=True)
    return out


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
